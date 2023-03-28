import openpyxl
import tkinter as tk

GREEN = "#1abc9c"
RED = "#e74c3c"
YELLOW = "#f1c40f"

class App:
    def __init__(self):
        self.window = tk.Tk()
        self.window.title("ICPC 2022")
        self.window.geometry("500x500")

        self.logo_img = tk.PhotoImage(file="icpc.png")

        self.logo_label = tk.Label(
            self.window,
            image=self.logo_img,
            bg="#34495e"
        )
        self.logo_label.place(x=10, y=20)

        self.bg_label = tk.Label(
            self.window,
            bg="#34495e",
            width=500,
            height=500
        )
        self.bg_label.place(x=0, y=0)

        self.title_label = tk.Label(
            self.window,
            text="ICPC Team Details",
            font=("Helvetica", 30),
            bg="#34495e",
            fg="white",
            highlightthickness=0
        )

        self.title_label.pack(pady=10)

        self.team_id_entry = tk.Entry(
            self.window,
            font=("Helvetica", 16),
            width=10
        )
        self.team_id_entry.pack(pady=10)

        self.search_btn = tk.Button(
            self.window,
            text="Search",
            font=("Helvetica", 16),
            bg="#2980b9",
            fg="white",
            command=self.search
        )
        self.search_btn.pack(pady=10)

        self.output_label = tk.Label(
            self.window,
            text="",
            font=("Helvetica", 16),
            bg="#34495e",
            fg="white",
            justify=tk.LEFT,
            anchor=tk.NW,
            wraplength=400
        )
        self.output_label.pack(pady=10)

        self.reset_btn = tk.Button(
            self.window,
            text="Reset",
            font=("Helvetica", 16),
            bg="#e74c3c",
            fg="white",
            command=self.reset
        )
        self.reset_btn.pack(pady=10)

        self.footer_label = tk.Label(
            self.window,
            text="coded by whitedevil",
            font=("Helvetica", 12),
            bg="#34495e",
            fg="white"
        )
        self.footer_label.pack(side=tk.BOTTOM, pady=20)

        self.wb = openpyxl.load_workbook("ICPC.xlsx")
        self.sh = self.wb.active

        self.c_values_printed = set()

    def run(self):
        self.window.mainloop()

    def search(self):
        user_input = self.team_id_entry.get()
        if user_input == "":
            return
        elif user_input == "000":
            self.window.destroy()
            return
        try:
            user_input = int(user_input)
        except ValueError:
            return
        if user_input == 0:
            return
        found = False
        output_text = ""
        for row in self.sh.iter_rows(min_row=2, max_row=self.sh.max_row, min_col=2, max_col=2):
            value_b = row[0].value
            if value_b == user_input:
                value_a = row[0].offset(column=-1).value
                value_c = row[0].offset(column=1).value
                if value_c not in self.c_values_printed:
                    output_text += f"\n\nTeam Name: {value_c}\n"
                    self.c_values_printed.add(value_c)
                    self.output_label.config(fg=GREEN)
                output_text += f"Team Member: {value_a}\n"
                found = True
                self.output_label.config(fg=YELLOW)
        if not found:
            output_text = "[-] ID Not Found"
            self.output_label.config(fg=RED)
        self.output_label.config(text=output_text)

    def reset(self):
        self.team_id_entry.delete(0, tk.END)
        self.output_label.config(text="")
        self.c_values_printed.clear()
        self.output_label.config(fg="white")

if __name__ == "__main__":
    app = App()
    app.run()
