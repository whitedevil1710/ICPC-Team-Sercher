import openpyxl


GREEN = '\033[32m'
RED = '\033[31m'
YELLOW = '\033[33m'
RESET = '\033[0m'

print('''
\033[38;2;255;0;0m██╗░█████╗░██████╗░░█████╗░  ██████╗░░█████╗░██████╗░██████╗░\033[0m
\033[38;2;255;140;0m██║██╔══██╗██╔══██╗██╔══██╗  ╚════██╗██╔══██╗╚════██╗╚════██╗\033[0m
\033[38;2;255;255;0m██║██║░░╚═╝██████╔╝██║░░╚═╝  ░░███╔═╝██║░░██║░░███╔═╝░░███╔═╝\033[0m
\033[38;2;0;255;0m██║██║░░██╗██╔═══╝░██║░░██╗  ██╔══╝░░██║░░██║██╔══╝░░██╔══╝░░\033[0m
\033[38;2;0;255;255m██║╚█████╔╝██║░░░░░╚█████╔╝  ███████╗╚█████╔╝███████╗███████╗\033[0m
\033[38;2;0;0;255m╚═╝░╚════╝░╚═╝░░░░░░╚════╝░  ╚══════╝░╚════╝░╚══════╝╚══════╝\033[0m
\033[38;2;255;255;255m-----------------c̶o̶d̶e̶d̶  b̶y̶  w̶h̶i̶t̶e̶d̶e̶v̶i̶l̶--------------------------\033[0m
''')


print("")
wb = openpyxl.load_workbook('ICPC.xlsx')
sh = wb.active

max_row = sh.max_row
user_input = ""

c_values_printed = set()

while user_input != 00:
    user_input = int(input("Enter the team ID (or 00 to quit): "))
    if user_input == 0:
        break
    found = False
    for row in sh.iter_rows(min_row=2, max_row=max_row, min_col=2, max_col=2):
        value_b = row[0].value
        if value_b == user_input:
            value_a = row[0].offset(column=-1).value
            value_c = row[0].offset(column=1).value
            if value_c not in c_values_printed:
                print(f"{GREEN}[+]Team Name: {value_c}{RESET}")
                c_values_printed.add(value_c)
            print(f"{YELLOW}[+]Team Member: {value_a}{RESET}")
            found = True
    if not found:
        print(f"{RED}[-]ID Not Found{RESET}")
